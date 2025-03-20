import cv2
import customtkinter
from tkinter import messagebox
from PIL import Image, ImageTk
from click import command
from customtkinter import CTkImage
import tkinter as tk
import webbrowser
import openpyxl
import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


# Configurações iniciais do CTkinter
customtkinter.set_appearance_mode("light")


def adicionar_dados_ao_excel(id_motorista, nome, sobrenome, idade, medicamentos, doencas, marca, modelo, ano, cor,
                             placa):
    # Caminho do arquivo
    caminho_arquivo = r"C:\Users\Pichau\Desktop\VISAO_COMPUTACIONAL\Dados coletados.xlsx"

    # Verifica se o arquivo existe
    if not os.path.exists(caminho_arquivo):
        # Se não existir, cria um novo arquivo
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Adiciona os cabeçalhos
        cabeçalhos = ["ID Motorista", "Nome", "Sobrenome", "Idade", "Medicamentos", "Doenças", "Marca", "Modelo", "Ano",
                      "Cor", "Placa"]
        sheet.append(cabeçalhos)
        workbook.save(caminho_arquivo)

    # Abre o arquivo existente
    workbook = openpyxl.load_workbook(caminho_arquivo)
    sheet = workbook.active

    # Encontra a primeira linha vazia
    linha = sheet.max_row + 1  # A próxima linha a ser preenchida
    while True:
        if sheet.cell(row=linha, column=1).value is None:  # Verifica a coluna A
            break
        linha += 1

    # Adiciona os dados nas colunas especificadas
    sheet.cell(row=linha, column=1, value=id_motorista)
    sheet.cell(row=linha, column=3, value=nome)
    sheet.cell(row=linha, column=4, value=sobrenome)
    sheet.cell(row=linha, column=5, value=idade)
    sheet.cell(row=linha, column=6, value=medicamentos)
    sheet.cell(row=linha, column=7, value=doencas)
    sheet.cell(row=linha, column=9, value=marca)
    sheet.cell(row=linha, column=10, value=modelo)
    sheet.cell(row=linha, column=11, value=ano)
    sheet.cell(row=linha, column=12, value=cor)
    sheet.cell(row=linha, column=13, value=placa)

    # Salva o arquivo
    workbook.save(caminho_arquivo)



class JanelaCapturaFoto(customtkinter.CTk):
    def __init__(self):
        super().__init__()
        self.geometry("600x700")
        self.title("Captura de Foto")
        self.attributes('-fullscreen', True)  # Tela cheia
        self.banner_path = "Banner Safe Vision Formulário (1200 x 800 px).png"
        self.banner_image = Image.open(self.banner_path)
        self.setup_ui()
        self.foto_capturada = None  # Inicializar variável para a foto capturada

    def redimensionar_banner(self):
        # Redimensiona a imagem do banner
        banner_width, banner_height = self.banner_image.size
        proporcao = banner_width / banner_height
        nova_largura = self.winfo_width()  # Largura atual da janela
        nova_altura = int(nova_largura / proporcao)  # Calcula nova altura mantendo a proporção
        return self.banner_image.resize((nova_largura, nova_altura), Image.LANCZOS)

    def setup_ui(self):
        customtkinter.CTkLabel(self, font=("Arial", 30), text="").pack(padx=10, pady=10)
        # Redimensiona o banner
        self.banner_redimensionado = self.redimensionar_banner()
        self.banner_tk = ImageTk.PhotoImage(self.banner_redimensionado)

        # Cria um label para exibir o banner
        banner_label = customtkinter.CTkLabel(self, image=self.banner_tk, text="")
        banner_label.pack(side="top", fill="x")  # Preenche horizontalmente

        customtkinter.CTkLabel(self, font=("Arial", 30), text="").pack(padx=10, pady=5)
        customtkinter.CTkLabel(self, font=("Arial", 20), text="Sistema para cadastro do motorista").pack(padx=10, pady=5)
        customtkinter.CTkLabel(self, font=("Arial", 20), text="Tire a sua foto para o cadastro").pack(padx=10, pady=5)
        customtkinter.CTkLabel(self, font=("Arial", 20), text="Olhe para a câmera e pressione o botão").pack(padx=10, pady=5)

        frame_borda = customtkinter.CTkFrame(self, width=410, height=410, fg_color="#0D6AD8")
        frame_borda.pack(pady=20)

        self.campo_foto = customtkinter.CTkLabel(frame_borda, width=400, height=400, text="", fg_color=self.cget("fg_color"), corner_radius=10)
        self.campo_foto.pack(padx=5, pady=5)

        customtkinter.CTkLabel(self, font=("Arial", 15), text="Insira abaixo o seu ID, exemplo: joao_silva.jpg").pack(padx=10, pady=0)
        self.id_entry = customtkinter.CTkEntry(self, width=300, placeholder_text="Digite aqui o seu ID")
        self.id_entry.pack(padx=10, pady=0)
        customtkinter.CTkLabel(self, font=("Arial", 15), text="Utilize o seu nome e sobrenome\nSomente letras minúsculas\nUtilize o _ entre o seu nome e sobrenome\nDeve conter .jpg no final").pack(padx=10, pady=0)
        customtkinter.CTkLabel(self, font=("Arial", 15), text="").pack(padx=10, pady=0)

        frame_botoes = customtkinter.CTkFrame(self, fg_color="transparent")
        frame_botoes.pack(pady=20)

        coluna1 = customtkinter.CTkFrame(frame_botoes, fg_color="transparent")
        coluna1.grid(row=0, column=0, padx=20)

        customtkinter.CTkButton(coluna1, text="Tirar Foto", command=self.tirar_foto, fg_color="#0D6AD8").pack(pady=10)
        customtkinter.CTkButton(coluna1, text="Salvar Foto e Dados", command=self.salvar_foto, fg_color="#0D6AD8").pack(pady=10)

        coluna2 = customtkinter.CTkFrame(frame_botoes, fg_color="transparent")
        coluna2.grid(row=0, column=1, padx=20)

        customtkinter.CTkButton(coluna2, text="Refazer Foto", command=self.refazer_foto, fg_color="#0D6AD8").pack(pady=10)
        customtkinter.CTkButton(coluna2, text="Finalizar", command=self.finalizar, fg_color="#0D6AD8").pack(pady=10)

    def tirar_foto(self):
        camera = cv2.VideoCapture(0)
        if not camera.isOpened():
            messagebox.showerror("Erro", "Não foi possível acessar a câmera.")
            return

        ret, frame = camera.read()
        if ret:
            self.foto_capturada = frame
            self.exibir_foto(frame)

        camera.release()

    def exibir_foto(self, frame):
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        imagem = Image.fromarray(frame_rgb)
        imagem_tk = ImageTk.PhotoImage(imagem)
        self.campo_foto.configure(image=imagem_tk)
        self.campo_foto.image = imagem_tk

    def salvar_foto(self):
        if self.foto_capturada is None:
            messagebox.showwarning("Aviso", "Nenhuma foto capturada.")
            return

        id_motorista = self.id_entry.get().strip()  # Corrigido para remover espaços desnecessários
        if not id_motorista or not id_motorista.endswith(".jpg"):
            messagebox.showwarning("Aviso", "Por favor, insira seu ID corretamente (ex: joao_silva.jpg).")
            return

        file_path = f"C:\\Users\\Pichau\\Desktop\\VISAO_COMPUTACIONAL\\Pessoas\\{id_motorista}"
        if not cv2.imwrite(file_path, self.foto_capturada):
            messagebox.showerror("Erro", "Erro ao salvar a imagem. Verifique o caminho ou permissões.")
            return

        messagebox.showinfo("Sucesso", f"Foto salva com êxito como '{id_motorista}'!")
        self.adicionar_dados_ao_excel(id_motorista)  # Salvar dados na planilha após salvar a foto

    def refazer_foto(self):
        self.campo_foto.configure(image="")
        self.foto_capturada = None

    def adicionar_dados_ao_excel(self, id_motorista):
        # Caminho do arquivo
        caminho_arquivo = r"C:\Users\Pichau\Desktop\VISAO_COMPUTACIONAL\Dados coletados.xlsx"

        # Verifica se o arquivo existe
        if not os.path.exists(caminho_arquivo):
            # Se não existir, cria um novo arquivo
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            # Adiciona os cabeçalhos
            cabeçalhos = ["ID Motorista", "Nome", "Sobrenome", "Idade", "Medicamentos", "Doenças", "Marca", "Modelo",
                          "Ano", "Cor", "Placa"]
            sheet.append(cabeçalhos)
            workbook.save(caminho_arquivo)

        # Abre o arquivo existente
        workbook = openpyxl.load_workbook(caminho_arquivo)
        sheet = workbook.active

        # Encontra a próxima linha vazia a partir da linha 3
        linha = 3
        while sheet.cell(row=linha, column=1).value is not None:
            linha += 1

        # Adiciona os dados nas colunas especificadas
        sheet.cell(row=linha, column=1, value=id_motorista)

        # Salva o arquivo
        workbook.save(caminho_arquivo)

    def finalizar(self):
        id_motorista = self.id_entry.get().strip()

        # Verifica se o ID do motorista não está vazio
        if not id_motorista:
            messagebox.showwarning("Aviso", "Por favor, insira seu ID antes de finalizar o cadastro.")
            return
        messagebox.showinfo("Fim", "Cadastro concluído.")
        self.destroy()

class JanelaInformacoesVeiculo(customtkinter.CTk):
    def __init__(self):
        super().__init__()
        self.geometry("600x700")
        self.title("Informações do veículo")
        self.attributes('-fullscreen', True)  # Tela cheia
        self.banner_path = "Banner Safe Vision Formulário (1200 x 800 px).png"
        self.banner_image = Image.open(self.banner_path)
        self.setup_ui()
        self.foto_capturada = None  # Inicializar variável para a foto capturada

    def redimensionar_banner(self):
        # Redimensiona a imagem do banner
        banner_width, banner_height = self.banner_image.size
        proporcao = banner_width / banner_height
        nova_largura = self.winfo_width()  # Largura atual da janela
        nova_altura = int(nova_largura / proporcao)  # Calcula nova altura mantendo a proporção
        return self.banner_image.resize((nova_largura, nova_altura), Image.LANCZOS)

    def setup_ui(self):
        customtkinter.CTkLabel(self, font=("Arial", 30), text="").pack(padx=10, pady=10)
        # Redimensiona o banner
        self.banner_redimensionado = self.redimensionar_banner()
        self.banner_tk = ImageTk.PhotoImage(self.banner_redimensionado)

        # Cria um label para exibir o banner
        banner_label = customtkinter.CTkLabel(self, image=self.banner_tk, text="")
        banner_label.pack(side="top", fill="x")  # Preenche horizontalmente

        customtkinter.CTkLabel(self, font=("Arial", 30), text="").pack(padx=10, pady=5)
        customtkinter.CTkLabel(self, font=("Arial", 30), text="Cadastro informações do veículo").pack(padx=10, pady=10)
        customtkinter.CTkLabel(self, font=("Arial", 15), text="").pack(padx=10, pady=10)

        # Criação das entradas
        customtkinter.CTkLabel(self, font=("Arial", 15), text="Insira abaixo a marca do seu veículo").pack(padx=10, pady=0)
        self.marca_entry = customtkinter.CTkEntry(self, width=300, placeholder_text="Coloque aqui a marca do veículo")
        self.marca_entry.pack(padx=10, pady=0)

        customtkinter.CTkLabel(self, font=("Arial", 15), text="Insira abaixo o modelo do veículo").pack(padx=10, pady=0)
        self.modelo_entry = customtkinter.CTkEntry(self, width=300, placeholder_text="Coloque aqui o modelo do veículo")
        self.modelo_entry.pack(padx=10, pady=0)

        customtkinter.CTkLabel(self, font=("Arial", 15), text="Insira abaixo o ano do veículo").pack(padx=10, pady=0)
        self.ano_entry = customtkinter.CTkEntry(self, width=300, placeholder_text="Coloque aqui o ano do veículo")
        self.ano_entry.pack(padx=10, pady=0)

        customtkinter.CTkLabel(self, font=("Arial", 15), text="Insira abaixo a cor do veículo").pack(padx=10, pady=0)
        self.cor_entry = customtkinter.CTkEntry(self, width=300, placeholder_text="Coloque aqui a cor do veículo")
        self.cor_entry.pack(padx=10, pady=0)

        customtkinter.CTkLabel(self, font=("Arial", 15), text="Insira abaixo a placa do veículo").pack(padx=10, pady=0)
        self.placa_entry = customtkinter.CTkEntry(self, width=300, placeholder_text="Coloque aqui a placa do veículo")
        self.placa_entry.pack(padx=10, pady=0)

        # Botão "Próximo"
        customtkinter.CTkButton(self, text="Salvar e Próxima etapa", command=self.salvar_dados_e_abrir_captura_foto,
                                fg_color="#0D6AD8").pack(pady=20)

    def adicionar_dados_ao_excel(self, marca, modelo, ano, cor, placa):
        # Caminho do arquivo
        caminho_arquivo = r"C:\Users\Pichau\Desktop\VISAO_COMPUTACIONAL\Dados coletados.xlsx"

        # Verifica se o arquivo existe
        if not os.path.exists(caminho_arquivo):
            # Se não existir, cria um novo arquivo
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            # Adiciona os cabeçalhos
            cabeçalhos = ["ID Motorista", "Nome", "Sobrenome", "Idade", "Medicamentos", "Doenças", "Marca", "Modelo",
                          "Ano", "Cor", "Placa"]
            sheet.append(cabeçalhos)
            workbook.save(caminho_arquivo)

        # Abre o arquivo existente
        workbook = openpyxl.load_workbook(caminho_arquivo)
        sheet = workbook.active

        # Encontra a próxima linha vazia a partir da linha 3
        linha = 3
        while sheet.cell(row=linha, column=9).value is not None:  # Verifica a coluna I (marca)
            linha += 1

        # Adiciona os dados nas colunas especificadas
        sheet.cell(row=linha, column=9, value=marca)  # Coluna I
        sheet.cell(row=linha, column=10, value=modelo)  # Coluna J
        sheet.cell(row=linha, column=11, value=ano)  # Coluna K
        sheet.cell(row=linha, column=12, value=cor)  # Coluna L
        sheet.cell(row=linha, column=13, value=placa)  # Coluna M

        # Salva o arquivo
        workbook.save(caminho_arquivo)

    def salvar_dados_e_abrir_captura_foto(self):
        # Salvar dados do veículo
        marca = self.marca_entry.get().strip()
        modelo = self.modelo_entry.get().strip()
        ano = self.ano_entry.get().strip()
        cor = self.cor_entry.get().strip()
        placa = self.placa_entry.get().strip()

        # Verifica se todos os campos estão preenchidos
        if not (marca and modelo and ano and cor and placa):
            messagebox.showwarning("Aviso", "Por favor, preencha todos os campos.")
            return

        # Adiciona os dados ao Excel
        self.adicionar_dados_ao_excel(marca, modelo, ano, cor, placa)

        # Mostra mensagem de sucesso
        messagebox.showinfo("Sucesso", "Dados do veículo salvos com sucesso!")

        self.destroy()  # Fechar a janela atual
        JanelaCapturaFoto().mainloop()  # Abrir a janela de captura de foto

# Janela inicial para informações do motorista
class JanelaInformacoesMotorista(customtkinter.CTk):
    def __init__(self):
        super().__init__()
        self.geometry("600x700")
        self.title("Informações do motorista")
        self.attributes('-fullscreen', True)  # Tela cheia
        self.banner_path = "Banner Safe Vision Formulário (1200 x 800 px).png"
        self.banner_image = Image.open(self.banner_path)
        self.setup_ui()
        self.foto_capturada = None  # Inicializar variável para a foto capturada

    def redimensionar_banner(self):
        # Redimensiona a imagem do banner
        banner_width, banner_height = self.banner_image.size
        proporcao = banner_width / banner_height
        nova_largura = self.winfo_width()  # Largura atual da janela
        nova_altura = int(nova_largura / proporcao)  # Calcula nova altura mantendo a proporção
        return self.banner_image.resize((nova_largura, nova_altura), Image.LANCZOS)

    def setup_ui(self):
        customtkinter.CTkLabel(self, font=("Arial", 30), text="").pack(padx=10, pady=10)
        # Redimensiona o banner
        self.banner_redimensionado = self.redimensionar_banner()
        self.banner_tk = ImageTk.PhotoImage(self.banner_redimensionado)

        # Cria um label para exibir o banner
        banner_label = customtkinter.CTkLabel(self, image=self.banner_tk, text="")
        banner_label.pack(side="top", fill="x")  # Preenche horizontalmente

        customtkinter.CTkLabel(self, font=("Arial", 30), text="").pack(padx=10, pady=5)
        customtkinter.CTkLabel(self, font=("Arial", 30), text="Cadastro informações do motorista").pack(padx=10, pady=10)
        customtkinter.CTkLabel(self, font=("Arial", 15), text="").pack(padx=10, pady=10)

        # Entradas para as informações do motorista
        customtkinter.CTkLabel(self, font=("Arial", 15), text="Insira abaixo o nome do motorista").pack(padx=10, pady=0)
        self.nome_entry = customtkinter.CTkEntry(self, width=300, placeholder_text="Coloque aqui o nome do motorista")
        self.nome_entry.pack(padx=10, pady=0)

        customtkinter.CTkLabel(self, font=("Arial", 15), text="Insira abaixo o sobrenome do motorista").pack(padx=10, pady=0)
        self.sobrenome_entry = customtkinter.CTkEntry(self, width=300, placeholder_text="Coloque aqui o sobrenome do motorista")
        self.sobrenome_entry.pack(padx=10, pady=0)

        customtkinter.CTkLabel(self, font=("Arial", 15), text="Insira abaixo a idade do motorista, somente números").pack(padx=10, pady=0)
        self.idade_entry = customtkinter.CTkEntry(self, width=300, placeholder_text="Coloque aqui a idade do motorista")
        self.idade_entry.pack(padx=10, pady=0)

        customtkinter.CTkLabel(self, font=("Arial", 15), text="Insira abaixo as informações sobre o uso de medicamentos").pack(padx=10, pady=0)
        self.medicamentos_entry = customtkinter.CTkEntry(self, width=300, placeholder_text="Utiliza algum medicamento? Se sim, qual?")
        self.medicamentos_entry.pack(padx=10, pady=0)

        customtkinter.CTkLabel(self, font=("Arial", 15), text="Insira abaixo se possui alguma doença").pack(padx=10, pady=0)
        self.doencas_entry = customtkinter.CTkEntry(self, width=300, placeholder_text="Tem alguma doença? Se sim, qual?")
        self.doencas_entry.pack(padx=10, pady=0)

        # Botão para salvar e ir para a próxima janela
        customtkinter.CTkButton(self, text="Salvar e Próxima Etapa", command=self.salvar_dados_e_abrir_janela_veiculo, fg_color="#0D6AD8").pack(pady=20)

    def adicionar_dados_ao_excel(self, nome, sobrenome, idade, medicamentos, doencas):
        # Caminho do arquivo
        caminho_arquivo = r"C:\Users\Pichau\Desktop\VISAO_COMPUTACIONAL\Dados coletados.xlsx"

        # Verifica se o arquivo existe
        if not os.path.exists(caminho_arquivo):
            # Se não existir, cria um novo arquivo
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            # Adiciona os cabeçalhos
            cabeçalhos = ["ID Motorista", "Nome", "Sobrenome", "Idade", "Medicamentos", "Doenças", "Marca", "Modelo",
                          "Ano", "Cor", "Placa"]
            sheet.append(cabeçalhos)
            workbook.save(caminho_arquivo)

        # Abre o arquivo existente
        workbook = openpyxl.load_workbook(caminho_arquivo)
        sheet = workbook.active

        # Encontra a primeira linha realmente vazia a partir da linha 3
        linha = 3  # Inicia a busca a partir da linha 3
        while not self.linha_vazia(sheet, linha):
            linha += 1

        # Adiciona os dados nas colunas correspondentes
        sheet.cell(row=linha, column=3, value=nome)  # Coluna C
        sheet.cell(row=linha, column=4, value=sobrenome)  # Coluna D
        sheet.cell(row=linha, column=5, value=idade)  # Coluna E
        sheet.cell(row=linha, column=6, value=medicamentos)  # Coluna F
        sheet.cell(row=linha, column=7, value=doencas)  # Coluna G

        # Salva o arquivo
        workbook.save(caminho_arquivo)

    def linha_vazia(self, sheet, linha):
        """Verifica se uma linha está completamente vazia."""
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=linha, column=col).value is not None:
                return False  # Se encontrar qualquer célula preenchida, a linha não está vazia
        return True  # Se todas as células da linha estiverem vazias, retorna True

    def salvar_dados_e_abrir_janela_veiculo(self):
        # Obtém os dados inseridos
        nome = self.nome_entry.get()
        sobrenome = self.sobrenome_entry.get()
        idade = self.idade_entry.get()
        medicamentos = self.medicamentos_entry.get()
        doencas = self.doencas_entry.get()

        # Salva os dados no Excel
        self.adicionar_dados_ao_excel(nome, sobrenome, idade, medicamentos, doencas)

        # Exibe uma mensagem de sucesso
        messagebox.showinfo("Sucesso", "Dados do motorista salvos com sucesso!")

        self.destroy()  # Fecha a janela atual
        JanelaInformacoesVeiculo().mainloop()  # Abre a janela de informações do veículo

customtkinter.set_appearance_mode("light")

class Tela_bem_vindo(customtkinter.CTk):
    def __init__(self):
        super().__init__()
        self.geometry("1200x800")
        self.title("Captura de Foto")
        self.attributes('-fullscreen', True)  # Tela cheia
        self.banner_path = "Banner Safe Vision Bem-vindo (1200 x 800 px).png"
        self.banner_image = Image.open(self.banner_path)
        self.setup_ui()
        self.foto_capturada = None  # Inicializar variável para a foto capturada

    def redimensionar_banner(self):
        # Mantém o tamanho original do banner
        return self.banner_image

    def setup_ui(self):
        # Configura o layout da tela
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)  # Permite o frame ocupar o espaço verticalmente

        # Redimensiona o banner
        self.banner_redimensionado = self.redimensionar_banner()
        self.banner_tk = ImageTk.PhotoImage(self.banner_redimensionado)

        # Cria um label para exibir o banner à esquerda
        banner_label = tk.Label(self, image=self.banner_tk)
        banner_label.grid(row=0, column=0, sticky="nsew")  # Banner ocupa a coluna da esquerda

        # Cria o frame para os botões e a mensagem à direita
        right_frame = customtkinter.CTkFrame(self, fg_color="transparent")  # Frame sem fundo
        right_frame.grid(row=0, column=1, padx=20, pady=20, sticky="nsew")

        # Configura o layout do right_frame
        right_frame.grid_rowconfigure(0, weight=1)  # Espaço antes da mensagem
        right_frame.grid_rowconfigure(1, weight=0)  # Mensagem
        right_frame.grid_rowconfigure(2, weight=0)  # Linha dos botões
        right_frame.grid_rowconfigure(3, weight=1)  # Espaço depois dos botões

        # Mensagem "Selecione uma das opções abaixo para prosseguir"
        info_text = customtkinter.CTkLabel(right_frame, font=("Arial", 25), text="Selecione uma das opções abaixo para prosseguir")
        info_text.grid(row=1, column=0, pady=20, padx=20)

        # Cria o frame para os botões abaixo da mensagem
        button_frame = customtkinter.CTkFrame(right_frame, fg_color="transparent")  # Remove o fundo
        button_frame.grid(row=2, column=0, pady=20, padx=20)

        # Botão "Iniciar Cadastro"
        iniciar_button = customtkinter.CTkButton(button_frame, text="Iniciar Cadastro", command=self.iniciar_cadastro, fg_color="#0D6AD8")
        iniciar_button.grid(row=0, column=0, pady=5)

        cadastrado_button = customtkinter.CTkButton(button_frame, text="Já tenho Cadastro", command=self.tem_cadastro, fg_color="#0D6AD8")
        cadastrado_button.grid(row=1, column=0, pady=5)

        # Botão "Encerrar Programa"
        encerrar_button = customtkinter.CTkButton(button_frame, text="Encerrar Programa", command=self.encerrar_programa, fg_color="#0D6AD8")
        encerrar_button.grid(row=2, column=0, pady=5)

        # Botão "Políticas de Privacidade e LGPD"
        privacidade_button = customtkinter.CTkButton(button_frame, text="Políticas de Privacidade e LGPD", command=self.abrir_politicas_privacidade, fg_color="#0D6AD8")
        privacidade_button.grid(row=3, column=0, pady=5)

    def iniciar_cadastro(self):
        self.destroy()  # Fecha a janela atual
        JanelaInformacoesMotorista().mainloop()  # Abre a janela de informações do veículo

    def tem_cadastro(self):
        messagebox.showinfo("Usuário já cadastrado", "Com o cadastro já realizado\nProssiga para o programa de detecção.")
        self.destroy()

    def encerrar_programa(self):
        # Função para fechar o programa
        self.destroy()

    def abrir_politicas_privacidade(self):
        # Função para abrir o site de Políticas de Privacidade e LGPD
        webbrowser.open("https://www.gov.br/anpd/pt-br/documentos-e-publicacoes/documentos-de-publicacoes/web-guia-anpd-tratamento-de-dados-para-fins-academicos.pdf")

# Execução da aplicação
if __name__ == "__main__":
    Tela_bem_vindo().mainloop()
